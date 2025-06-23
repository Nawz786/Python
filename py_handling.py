from openpyxl import Workbook, load_workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Data"
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet["C1"] = "Gender"
sheet.append(["Mike", 24, "Male"])
sheet.append(["Nike", 23, "Male"])
file_path = "/Users/nawaz/Desktop/Py_sample.xlsx"
wb.save(file_path)
wb = load_workbook(file_path)
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    print(row)

-------
from bs4 import BeautifulSoup

xml_data = """
<books>
    <book id="1">
        <title>The Great Gatsby</title>
        <author>Harvey</author>
    </book>
</books>
"""
file_path = "/Users/nawaz/Desktop/py/sample_x.xml"
soup = BeautifulSoup(xml_data, "lxml-xml")
with open(file_path, "w") as file:
    file.write(soup.prettify())
with open(file_path, "r") as file:
    soup = BeautifulSoup(file, "lxml-xml")
    author = soup.find("author").text
    print(author)
--------

import json
data = {
    "person": {
        "name": "John",
        "age": 30
    },
    "student": {
        "name": "Alice",
        "age": 25
    }
}
file_path = "/Users/nawaz/Desktop/py/sample_json_file.json"
with open(file_path, "w") as file:
    json.dump(data, file, indent=2)
with open(file_path, "r") as file:
    data = json.load(file)
print(data["student"]["name"])
print(data["person"])

---------

file_path = "/Users/nawaz/Desktop/py/sample.txt"
with open(file_path, "w") as file:
    file.write("file handling")
with open(file_path, "r") as file:
    content = file.read()
    print(content)
with open(file_path, "a") as file:
    file.write("\nThis is an addition line to the file")
